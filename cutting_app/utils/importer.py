import csv
import openpyxl
from typing import List


class Importer:
    def import_csv(self, file_path: str) -> List[dict]:
        items = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                item = {
                    "name": row.get("name", row.get("Название", row.get("Наименование", "Unknown"))),
                    "width_mm": float(row.get("width", row.get("Ширина", row.get("width_mm", 0)))),
                    "height_mm": float(row.get("height", row.get("Высота", row.get("height_mm", 0)))),
                    "quantity": int(row.get("quantity", row.get("Количество", row.get("qty", 1)))),
                    "rotation": row.get("rotation", row.get("Вращение", "true")).lower() in ("true", "yes", "1", "да"),
                    "fibers": row.get("fibers", row.get("Волокна", "any")),
                    "material_type_id": int(row.get("material_type_id", row.get("material", 1)))
                }
                items.append(item)
        return items

    def import_excel(self, file_path: str) -> List[dict]:
        items = []
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        name_cols = ["name", "название", "наименование"]
        width_cols = ["width", "ширина", "width_mm"]
        height_cols = ["height", "высота", "height_mm"]
        qty_cols = ["quantity", "количество", "qty", "кол-во"]

        name_idx = next((i for i, h in enumerate(headers) if h and any(nc in str(h).lower() for nc in name_cols)), 0)
        width_idx = next((i for i, h in enumerate(headers) if h and any(wc in str(h).lower() for wc in width_cols)), 1)
        height_idx = next((i for i, h in enumerate(headers) if h and any(hc in str(h).lower() for hc in height_cols)), 2)
        qty_idx = next((i for i, h in enumerate(headers) if h and any(qc in str(h).lower() for qc in qty_cols)), 3)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[name_idx]:
                continue

            item = {
                "name": str(row[name_idx]),
                "width_mm": float(row[width_idx]) if row[width_idx] else 0,
                "height_mm": float(row[height_idx]) if row[height_idx] else 0,
                "quantity": int(row[qty_idx]) if row[qty_idx] else 1,
                "rotation": True,
                "fibers": "any",
                "material_type_id": 1
            }
            items.append(item)

        return items