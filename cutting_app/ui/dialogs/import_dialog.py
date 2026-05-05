from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QFileDialog, QProgressBar, QTableWidget)
from PyQt6.QtCore import Qt
import csv
import openpyxl


class ImportDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Импорт спецификации")
        self.setMinimumWidth(500)
        self._setup_ui()
        self.file_path = None
        self.items = []

    def _setup_ui(self):
        layout = QVBoxLayout(self)

        self.info_label = QLabel("Выберите файл CSV или Excel (.xlsx) для импорта спецификации")
        layout.addWidget(self.info_label)

        file_layout = QHBoxLayout()
        self.file_label = QLabel("Файл не выбран")
        self.btn_browse = QPushButton("Обзор...")
        self.btn_browse.clicked.connect(self._browse_file)
        file_layout.addWidget(self.file_label, 1)
        file_layout.addWidget(self.btn_browse)
        layout.addLayout(file_layout)

        self.btn_import_csv = QPushButton("Импорт CSV")
        self.btn_import_csv.clicked.connect(self._import_csv)
        self.btn_import_xlsx = QPushButton("Импорт Excel")
        self.btn_import_xlsx.clicked.connect(self._import_xlsx)

        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.btn_import_csv)
        btn_layout.addWidget(self.btn_import_xlsx)
        btn_layout.addStretch()
        self.btn_ok = QPushButton("ОК")
        self.btn_ok.clicked.connect(self.accept)
        self.btn_ok.setEnabled(False)
        btn_layout.addWidget(self.btn_ok)
        btn_cancel = QPushButton("Отмена")
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_cancel)

        layout.addLayout(btn_layout)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(6)
        self.preview_table.setHorizontalHeaderLabels(["Название", "Ширина", "Высота", "Кол-во", "Вращение", "Волокна"])
        layout.addWidget(self.preview_table)

        self.status_label = QLabel("")
        layout.addWidget(self.status_label)

    def _browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл", "", 
            "CSV Files (*.csv);;Excel Files (*.xlsx);;All Files (*.*)"
        )
        if file_path:
            self.file_path = file_path
            self.file_label.setText(file_path.split("/")[-1].split("\\")[-1])

    def _import_csv(self):
        if not self.file_path:
            self.status_label.setText("Сначала выберите файл")
            return

        try:
            self.progress.setVisible(True)
            self.progress.setRange(0, 0)
            self.items = []

            with open(self.file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames
                
                name_col = self._find_column(headers, ["name", "название", "наименование", "деталь"])
                width_col = self._find_column(headers, ["width", "ширина", "width_mm"])
                height_col = self._find_column(headers, ["height", "высота", "height_mm"])
                qty_col = self._find_column(headers, ["quantity", "количество", "qty", "кол-во"])
                rotation_col = self._find_column(headers, ["rotation", "вращение", "поворот"])
                fibers_col = self._find_column(headers, ["fibers", "волокна"])

                for row in reader:
                    name = row.get(name_col, "Unknown") if name_col else "Unknown"
                    width = float(row.get(width_col, 0)) if width_col and row.get(width_col) else 0
                    height = float(row.get(height_col, 0)) if height_col and row.get(height_col) else 0
                    qty = int(row.get(qty_col, 1)) if qty_col and row.get(qty_col) else 1
                    rotation = row.get(rotation_col, "true").lower() in ("true", "yes", "1", "да") if rotation_col else True
                    fibers = row.get(fibers_col, "any") if fibers_col else "any"

                    if width > 0 and height > 0:
                        self.items.append({
                            "name": name,
                            "width_mm": width,
                            "height_mm": height,
                            "quantity": qty,
                            "rotation": rotation,
                            "fibers": fibers,
                            "material_type_id": 1
                        })

            self._show_preview()
            self.status_label.setText(f"Импортировано {len(self.items)} позиций")

        except Exception as e:
            self.status_label.setText(f"Ошибка: {str(e)}")
        finally:
            self.progress.setVisible(False)

    def _import_xlsx(self):
        if not self.file_path:
            self.status_label.setText("Сначала выберите файл")
            return

        try:
            self.progress.setVisible(True)
            self.progress.setRange(0, 0)
            self.items = []

            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            headers = [str(cell.value).lower() if cell.value else "" for cell in ws[1]]

            name_idx = self._find_index(headers, ["name", "название", "наименование", "деталь"])
            width_idx = self._find_index(headers, ["width", "ширина", "width_mm"])
            height_idx = self._find_index(headers, ["height", "высота", "height_mm"])
            qty_idx = self._find_index(headers, ["quantity", "количество", "qty", "кол-во"])
            rotation_idx = self._find_index(headers, ["rotation", "вращение"])
            fibers_idx = self._find_index(headers, ["fibers", "волокна"])

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[name_idx] if name_idx >= 0 else True:
                    continue

                name = str(row[name_idx]) if name_idx >= 0 else "Unknown"
                width = float(row[width_idx]) if width_idx >= 0 and row[width_idx] else 0
                height = float(row[height_idx]) if height_idx >= 0 and row[height_idx] else 0
                qty = int(row[qty_idx]) if qty_idx >= 0 and row[qty_idx] else 1
                rotation = str(row[rotation_idx]).lower() in ("true", "yes", "1", "да") if rotation_idx >= 0 else True
                fibers = str(row[fibers_idx]) if fibers_idx >= 0 else "any"

                if width > 0 and height > 0:
                    self.items.append({
                        "name": name,
                        "width_mm": width,
                        "height_mm": height,
                        "quantity": qty,
                        "rotation": rotation,
                        "fibers": fibers,
                        "material_type_id": 1
                    })

            self._show_preview()
            self.status_label.setText(f"Импортировано {len(self.items)} позиций")

        except Exception as e:
            self.status_label.setText(f"Ошибка: {str(e)}")
        finally:
            self.progress.setVisible(False)

    def _show_preview(self):
        self.preview_table.setRowCount(len(self.items))
        for i, item in enumerate(self.items):
            self.preview_table.setItem(i, 0, item["name"])
            self.preview_table.setItem(i, 1, str(item["width_mm"]))
            self.preview_table.setItem(i, 2, str(item["height_mm"]))
            self.preview_table.setItem(i, 3, str(item["quantity"]))
            self.preview_table.setItem(i, 4, "Да" if item["rotation"] else "Нет")
            self.preview_table.setItem(i, 5, item["fibers"])
        self.preview_table.resizeColumnsToContents()
        self.btn_ok.setEnabled(len(self.items) > 0)

    def _find_column(self, headers: list, variants: list) -> str:
        for h in headers:
            h_lower = str(h).lower()
            for v in variants:
                if v in h_lower:
                    return h
        return None

    def _find_index(self, headers: list, variants: list) -> int:
        for i, h in enumerate(headers):
            h_lower = str(h).lower()
            for v in variants:
                if v in h_lower:
                    return i
        return -1

    def get_items(self):
        return self.items